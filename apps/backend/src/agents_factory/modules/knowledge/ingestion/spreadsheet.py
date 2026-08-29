from __future__ import annotations

import json
from io import BytesIO

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.utils.exceptions import InvalidFileException  # type: ignore[import-untyped]

from agents_factory.modules.knowledge.ingestion.contracts import (
    ExtractedBlock,
    ExtractedDocument,
    FetchedSource,
    IngestionRejected,
)


class SpreadsheetExtractor:
    def __init__(self, *, max_rows: int = 50_000, max_columns: int = 200) -> None:
        self._max_rows = max_rows
        self._max_columns = max_columns

    def extract(self, fetched: FetchedSource) -> ExtractedDocument:
        if fetched.media_type == "application/json":
            return self._extract_google_sheet_rows(fetched)
        try:
            workbook = load_workbook(
                BytesIO(fetched.content), read_only=True, data_only=True
            )
        except (InvalidFileException, OSError, ValueError):
            raise IngestionRejected("spreadsheet_malformed") from None
        blocks: list[ExtractedBlock] = []
        total_rows = 0
        for sheet in workbook.worksheets:
            if sheet.max_column > self._max_columns:
                raise IngestionRejected("spreadsheet_column_limit_exceeded")
            rows: list[tuple[str, ...]] = []
            for row in sheet.iter_rows(values_only=True):
                total_rows += 1
                if total_rows > self._max_rows:
                    raise IngestionRejected("spreadsheet_row_limit_exceeded")
                values = tuple(
                    "" if cell is None else str(cell).strip() for cell in row
                )
                if any(values):
                    rows.append(values)
            if rows:
                blocks.append(
                    ExtractedBlock(
                        kind="TABLE",
                        text="\n".join(" | ".join(row) for row in rows),
                        rows=tuple(rows),
                        locator={**fetched.locator, "sheet": sheet.title},
                    )
                )
        workbook.close()
        return self._document(fetched, blocks)

    def _extract_google_sheet_rows(self, fetched: FetchedSource) -> ExtractedDocument:
        try:
            payload = json.loads(fetched.content)
        except (UnicodeDecodeError, json.JSONDecodeError):
            raise IngestionRejected("google_sheet_rows_malformed") from None
        if not isinstance(payload, dict) or set(payload) != {"sheet", "rows"}:
            raise IngestionRejected("google_sheet_rows_malformed")
        sheet = payload["sheet"]
        raw_rows = payload["rows"]
        if not isinstance(sheet, str) or not isinstance(raw_rows, list):
            raise IngestionRejected("google_sheet_rows_malformed")
        if len(raw_rows) > self._max_rows:
            raise IngestionRejected("spreadsheet_row_limit_exceeded")
        rows: list[tuple[str, ...]] = []
        for raw_row in raw_rows:
            if not isinstance(raw_row, list) or len(raw_row) > self._max_columns:
                raise IngestionRejected("google_sheet_rows_malformed")
            rows.append(tuple(str(cell) for cell in raw_row))
        block = ExtractedBlock(
            kind="TABLE",
            text="\n".join(" | ".join(row) for row in rows),
            rows=tuple(rows),
            locator={**fetched.locator, "sheet": sheet},
        )
        return self._document(fetched, [block])

    @staticmethod
    def _document(
        fetched: FetchedSource, blocks: list[ExtractedBlock]
    ) -> ExtractedDocument:
        if not blocks:
            raise IngestionRejected("source_has_no_extractable_text")
        return ExtractedDocument(
            title=fetched.filename,
            blocks=tuple(blocks),
            source_digest=fetched.content_digest,
        )
